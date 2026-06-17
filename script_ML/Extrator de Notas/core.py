import csv
import io
import json
import time
import re
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import COOKIE_PATH, JSON_EXTRATO, JSON_ENRIQ, EXCEL_CLASS, XML_DIR, JSON_FINAL
import auth_ml

REQUEST_TIMEOUT = 15  # segundos — evita que uma request travada pendure o script inteiro

# ── Helpers de sessão / csrf / busca direta no frontend ──────────────────────

def _carregar_sessao() -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})
    try:
        cookies = json.loads(COOKIE_PATH.read_text(encoding='utf-8'))
        for c in cookies:
            session.cookies.set(c['name'], c['value'], domain=c['domain'])
    except Exception as e:
        raise RuntimeError(f"Erro ao ler cookie.json: {e}")
    return session


def _obter_csrf_token(session: requests.Session, log_cb):
    """Pega o csrf-token da meta tag da página, necessário pra busca direta no list_items."""
    log_cb("Obtendo csrf-token da sessão...")
    try:
        r = session.get(
            "https://myaccount.mercadolivre.com.br/purchases/list",
            timeout=REQUEST_TIMEOUT,
        )
        m = re.search(r'name="csrf-token"\s+content="([^"]+)"', r.text)
        if m:
            return m.group(1)
    except Exception:
        pass
    log_cb("AVISO: csrf-token não encontrado. Busca direta de compra ficará indisponível (usará pack_id como fallback).")
    return None


def _achar_contexto(obj, valor_busca: str):
    """Percorre recursivamente a resposta JSON do list_items procurando o bloco
    'context' (purchase_id/order_id/pack_id) que corresponde ao valor buscado."""
    if isinstance(obj, dict):
        ctx = obj.get("context")
        if isinstance(ctx, dict) and valor_busca in (
            str(ctx.get("purchase_id")), str(ctx.get("order_id")), str(ctx.get("pack_id"))
        ):
            return ctx
        for v in obj.values():
            achado = _achar_contexto(v, valor_busca)
            if achado:
                return achado
    elif isinstance(obj, list):
        for item in obj:
            achado = _achar_contexto(item, valor_busca)
            if achado:
                return achado
    return None


def _buscar_contexto_compra(session: requests.Session, csrf_token, valor_busca: str):
    """Busca a compra direto pelo order_id/pack_id no endpoint de busca do my_purchases,
    evitando ter que varrer o histórico inteiro em 20 páginas fixas."""
    if not csrf_token or not valor_busca:
        return None
    params = {
        "requestInfo[method]": "GET",
        "requestInfo[path]": "/my_purchases/api/web/list_items",
        "requestInfo[params][0][key]": "searchValue",
        "requestInfo[params][0][value]": valor_busca,
    }
    headers = {
        "x-csrf-token": csrf_token,
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://myaccount.mercadolivre.com.br/my_purchases/list",
    }
    try:
        r = session.get(
            "https://myaccount.mercadolivre.com.br/my_purchases/api/web/list_items",
            params=params, headers=headers, timeout=REQUEST_TIMEOUT,
        )
        if r.status_code != 200:
            return None
        return _achar_contexto(r.json(), valor_busca)
    except Exception:
        return None


def _buscar_frete_csv(session: requests.Session, csrf_token, order_id: str, client_id: str, log_cb) -> float:
    """Busca o custo de envio real do pedido via endpoint download-csv do frontend.
    Retorna o valor numérico da coluna 'Custo de envio' para o order_id informado,
    ou 0.0 se não encontrado / em caso de erro."""
    if not csrf_token or not order_id:
        return 0.0
    params = {
        "requestInfo[method]": "GET",
        "requestInfo[path]": "/my_purchases/middleend/web/report/csv",
        "requestInfo[params][0][key]": "client.id",
        "requestInfo[params][0][value]": client_id,
        "requestInfo[params][1][key]": "device.type",
        "requestInfo[params][1][value]": "desktop",
        "requestInfo[params][2][key]": "search.value",
        "requestInfo[params][2][value]": order_id,
    }
    headers = {
        "x-csrf-token": csrf_token,
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://myaccount.mercadolivre.com.br/my_purchases/list",
    }
    try:
        r = session.get(
            "https://myaccount.mercadolivre.com.br/my_purchases/api/web/download-csv",
            params=params, headers=headers, timeout=REQUEST_TIMEOUT,
        )
        if r.status_code != 200:
            return 0.0
        # Decodifica com utf-8-sig para remover BOM (\ufeff) se presente,
        # e cai em latin-1 se utf-8 falhar (encoding comum em CSVs do ML)
        try:
            texto = r.content.decode("utf-8-sig")
        except UnicodeDecodeError:
            texto = r.content.decode("latin-1")
        reader = csv.DictReader(io.StringIO(texto), delimiter=";")
        for row in reader:
            pedido_col = row.get("N.º do pedido", "").strip()
            if pedido_col == order_id:
                raw = row.get("Custo de envio", "0").strip().replace(",", ".")
                try:
                    return float(raw)
                except ValueError:
                    return 0.0
    except Exception as e:
        log_cb(f"  AVISO: erro ao buscar frete CSV para {order_id}: {e}")
    return 0.0


# ── 1. EXTRAÇÃO ──────────────────────────────────────────────────────────────
def executar_extracao(cfg: dict, auth: dict, dt_inicio: str, dt_fim: str, log_cb) -> dict:
    auth = auth_ml.ensure_valid_token(cfg, auth)

    log_cb("Carregando sessão (cookies)...")
    session = _carregar_sessao()
    csrf_token = _obter_csrf_token(session, log_cb)
    client_id = str(auth.get("seller_id", ""))

    # Cache incremental: reaproveita pedidos já extraídos em execuções anteriores,
    # sem repetir scraping/HTTP pra quem já está pronto.
    cache_existente = {}
    if JSON_EXTRATO.exists():
        try:
            for p in json.loads(JSON_EXTRATO.read_text(encoding="utf-8")):
                cache_existente[str(p.get("pedido_id"))] = p
        except Exception:
            cache_existente = {}

    log_cb("Buscando pedidos pagos via API...")
    pedidos_api, offset, limit = [], 0, 50
    headers = {"Authorization": f"Bearer {auth['access_token']}"}
    while True:
        params = {
            "buyer": auth["seller_id"],
            "order.date_created.from": f"{dt_inicio}T00:00:00.000-03:00",
            "order.date_created.to": f"{dt_fim}T23:59:59.000-03:00",
            "order.status": "paid", "offset": offset, "limit": limit, "sort": "date_desc"
        }
        r = session.get(
            "https://api.mercadolibre.com/orders/search",
            headers=headers, params=params, timeout=REQUEST_TIMEOUT
        )
        r.raise_for_status()
        data = r.json()
        resultados = data.get("results", [])
        pedidos_api.extend(resultados)
        if offset + limit >= data.get("paging", {}).get("total", 0) or not resultados:
            break
        offset += limit
        time.sleep(0.3)

    n = len(pedidos_api)
    log_cb(f"{n} pedido(s) encontrado(s).")
    if n == 0:
        return auth

    json_erp = []
    XML_DIR.mkdir(parents=True, exist_ok=True)
    reaproveitados = 0

    for i, order in enumerate(pedidos_api, 1):
        o_id = str(order.get("id"))

        # já processado numa execução anterior: reaproveita sem bater na rede de novo
        if o_id in cache_existente:
            json_erp.append(cache_existente[o_id])
            reaproveitados += 1
            continue

        pk_id = str(order.get("pack_id")) if order.get("pack_id") else o_id
        log_cb(f"[{i:03d}/{n:03d}] Processando pedido {o_id}...")

        # Cálculo de parcelas e diff de juros/desconto (ajuste final montado após calcular frete)
        pags = order.get("payments", [])
        pag = next((p for p in pags if p.get("status") == "approved"), pags[0] if pags else {})
        parcelas = pag.get("installments", 1)
        diff = round(pag.get("total_paid_amount", 0.0) - pag.get("transaction_amount", 0.0), 2)

        # Busca direta da compra no frontend (substitui a varredura fixa de 20 páginas)
        ctx = _buscar_contexto_compra(session, csrf_token, o_id) or _buscar_contexto_compra(session, csrf_token, pk_id)
        purch_id = (ctx.get("purchase_id") if ctx else None) or pk_id

        # Raspagem Frontend (Cartão)
        cartao_real = "N/A"
        try:
            html = session.get(
                f"https://myaccount.mercadolivre.com.br/my_purchases/{purch_id}/status?packId={pk_id}&orderId={o_id}",
                timeout=REQUEST_TIMEOUT
            ).text
            m = re.search(r'"cardNameAndLastDigits"\s*:\s*"([^"]+)"', html)
            if m: cartao_real = m.group(1).strip()
            elif '"payment_method_id":"pix"' in html: cartao_real = "PIX"
            elif '"payment_method_id":"account_money"' in html: cartao_real = "Saldo em Conta"
        except Exception:
            pass

        # XML: só bate no invoices-overview se ainda não tiver o arquivo em disco
        caminho_xml = "N/A"
        xml_path = XML_DIR / f"nfe_{o_id}.xml"
        if xml_path.exists():
            caminho_xml = str(xml_path.resolve())
        else:
            try:
                r_ov = session.get(
                    f"https://www.mercadolivre.com.br/emissor/omni/api/invoices-overview?identifiers={o_id}",
                    timeout=REQUEST_TIMEOUT
                )
                xml_url = next(
                    (sub.get("url") for inv in r_ov.json().get("invoices", [])
                     for act in inv.get("actions", [])
                     for sub in act.get("sub_actions", [])
                     if sub.get("id") == "download_xml"),
                    None
                )
                if xml_url:
                    r_xml = session.get(xml_url, timeout=REQUEST_TIMEOUT)
                    if r_xml.status_code == 200:
                        xml_path.write_bytes(r_xml.content)
                        caminho_xml = str(xml_path.resolve())
            except Exception:
                pass

        # Frete: usa a API como prioridade. Busca no CSV apenas se a API retornar 0
        frete_api = float(order.get("shipping_cost") or pag.get("shipping_cost", 0.0))
        
        if frete_api == 0.0:
            frete_csv = _buscar_frete_csv(session, csrf_token, o_id, client_id, log_cb)
            frete = frete_csv if frete_csv > 0.0 else 0.0
        else:
            frete = frete_api

        # Ajuste financeiro: combina juros/desconto (diff) com frete
        tem_juros = diff > 0
        tem_frete = frete > 0.0
        if tem_juros and tem_frete:
            ajuste = {"tipo": "juros + frete", "valor": round(diff + frete, 2)}
        elif tem_juros:
            ajuste = {"tipo": "juros", "valor": round(diff, 2)}
        elif diff < 0:
            ajuste = {"tipo": "desconto", "valor": round(abs(diff), 2)}
        elif tem_frete:
            ajuste = {"tipo": "frete", "valor": round(frete, 2)}
        else:
            ajuste = {"tipo": "nenhum", "valor": 0.0}
        v_total = float(pag.get("total_paid_amount", order.get("total_amount", 0.0)))
        dt_criacao = order.get("date_created", "")[:10]

        # Título e id_pagamento: usa o que já vem do /orders/search.
        # Se faltar, o Passo 2 (enriquecimento) completa de forma lazy — não tem
        # mais a busca completa forçada pra todo pedido aqui no Passo 1.
        id_pag = order.get("buyer", {}).get("billing_info", {}).get("id")
        titulos = [it.get("item", {}).get("title", "Desconhecido") for it in order.get("order_items", [])]
        nome_produto_real = " | ".join(titulos) if titulos else "N/A"

        obj = {
            "pedido_id": o_id,
            "id_pagamento": id_pag,
            "nome_produto": nome_produto_real,
            "data_compra": dt_criacao,
            "valor_total": v_total,
            "cartao_utilizado": cartao_real,
            "parcelas": parcelas,
            "frete": frete,
            "ajuste_financeiro": ajuste,
            "caminho_xml": caminho_xml,
            "dados_manuais": None
        }

        if caminho_xml == "N/A":
            obj["dados_manuais"] = {
                "nome_produto": nome_produto_real,
                "data_compra": dt_criacao,
                "valor_total": v_total,
                "valor_unitario_parcela": round(v_total / parcelas, 2) if parcelas > 0 else v_total
            }
        json_erp.append(obj)
        time.sleep(0.3)

    if reaproveitados:
        log_cb(f"{reaproveitados} pedido(s) reaproveitado(s) do cache (sem nova requisição).")

    JSON_EXTRATO.write_text(json.dumps(json_erp, ensure_ascii=False, indent=2), encoding="utf-8")
    log_cb("✓ Arquivo base integracao_erp.json salvo.")
    return auth

# ── 2. ENRIQUECIMENTO ────────────────────────────────────────────────────────

def executar_enriquecimento(cfg: dict, auth: dict, log_cb) -> dict:
    if not JSON_EXTRATO.exists():
        raise RuntimeError("Execute a extração (Passo 1) primeiro.")

    auth = auth_ml.ensure_valid_token(cfg, auth)
    session = requests.Session()
    session.headers.update({"Authorization": f"Bearer {auth['access_token']}"})
    pedidos = json.loads(JSON_EXTRATO.read_text(encoding="utf-8"))

    log_cb(f"Enriquecendo {len(pedidos)} pedidos com títulos da API...")
    for i, p in enumerate(pedidos, 1):
        oid = p["pedido_id"]
        precisa_titulo = not p.get("nome_produto") or p["nome_produto"] == "N/A"
        precisa_id_pag = not p.get("id_pagamento")

        if precisa_titulo or precisa_id_pag:
            log_cb(f"[{i:03d}/{len(pedidos):03d}] Buscando dados complementares para {oid}...")
            try:
                resp = session.get(f"https://api.mercadolibre.com/orders/{oid}", timeout=REQUEST_TIMEOUT)
            except requests.RequestException:
                resp = None
            if resp is not None and resp.status_code == 200:
                order_data = resp.json()
                if precisa_titulo:
                    titulos = [it.get("item", {}).get("title", "Desconhecido") for it in order_data.get("order_items", [])]
                    p["nome_produto"] = " | ".join(titulos) if titulos else "N/A"
                if precisa_id_pag:
                    p["id_pagamento"] = order_data.get("buyer", {}).get("billing_info", {}).get("id")
            else:
                if precisa_titulo: p["nome_produto"] = "N/A"
            time.sleep(0.3)

    JSON_ENRIQ.write_text(json.dumps(pedidos, ensure_ascii=False, indent=2), encoding="utf-8")
    log_cb("✓ Arquivo integracao_erp_enriquecido.json salvo.")

    # Exportar Excel (lógica de colunas inalterada)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classificação ERP"
    colunas = [
        ("pedido_id", "Pedido ID", 22),
        ("data_compra", "Data", 12),
        ("nome_produto", "Nome do Produto", 50),
        ("conta_resultado", "Conta Resultado", 28),
        ("cartao_utilizado", "Cartão", 24),
        ("parcelas", "Parcelas", 10),
        ("valor_total", "Valor (R$)", 14),
        ("id_pagamento", "ID do Pagamento", 20),
        ("frete", "Frete (R$)", 14),
        ("ajuste_tipo", "Ajuste Tipo", 14),
        ("ajuste_valor", "Ajuste (R$)", 18),
        ("tem_xml", "Tem NFe (XML)?", 16)
    ]

    for c_idx, (_, lbl, w) in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=c_idx, value=lbl)
        cell.font = Font(bold=True, color="F5A623")
        cell.fill = PatternFill("solid", fgColor="1A1D27")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w

    for r_idx, p in enumerate(pedidos, 2):
        ajuste = p.get("ajuste_financeiro", {})
        tem_xml = "Sim" if p.get("caminho_xml") and p.get("caminho_xml") != "N/A" else "Não"

        linha = [
            p.get("pedido_id", ""),
            p.get("data_compra", ""),
            p.get("nome_produto", ""),
            "",
            p.get("cartao_utilizado", ""),
            p.get("parcelas", 1),
            p.get("valor_total", 0.0),
            p.get("id_pagamento", ""),
            p.get("frete", 0.0),
            ajuste.get("tipo", ""),
            ajuste.get("valor", 0.0),
            tem_xml
        ]
        for c_idx, val in enumerate(linha, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if c_idx == 4:
                cell.fill = PatternFill("solid", fgColor="FFF8E7")
                cell.font = Font(bold=True, color="B8860B")

    wb.save(EXCEL_CLASS)
    log_cb(f"✓ Planilha gerada: {EXCEL_CLASS.name}. Preencha-a antes do Passo 3!")
    return auth

# ── 3. CLASSIFICAÇÃO FINAL ───────────────────────────────────────────────────

def executar_classificacao(log_cb):
    if not JSON_ENRIQ.exists() or not EXCEL_CLASS.exists():
        raise RuntimeError("Arquivos ausentes. Execute o Passo 2 e preencha a planilha.")

    log_cb("Lendo planilha de classificação...")
    wb = openpyxl.load_workbook(EXCEL_CLASS, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    idx_id = next((i for i, h in enumerate(header) if "pedido" in h), None)
    idx_cr = next((i for i, h in enumerate(header) if "conta resultado" in h), None)

    if idx_id is None or idx_cr is None:
        raise RuntimeError("Colunas 'Pedido ID' ou 'Conta Resultado' não encontradas no Excel.")

    mapa = {str(r[idx_id]).strip(): str(r[idx_cr]).strip() for r in rows[1:] if r[idx_id] and str(r[idx_id]).strip() != "None"}

    pedidos = json.loads(JSON_ENRIQ.read_text(encoding="utf-8"))
    sem_class = []

    for p in pedidos:
        oid = p["pedido_id"]
        conta = mapa.get(oid, "")
        if not conta or conta == "None":
            sem_class.append(oid)

        novo = {}
        for k, v in p.items():
            novo[k] = v
            if k == "nome_produto": novo["contaResultados"] = conta
        if "contaResultados" not in novo: novo["contaResultados"] = conta
        p.clear()
        p.update(novo)

    JSON_FINAL.write_text(json.dumps(pedidos, ensure_ascii=False, indent=2), encoding="utf-8")

    if sem_class:
        log_cb(f"AVISO: {len(sem_class)} pedido(s) sem classificação detectados.")
    log_cb(f"✓ JSON Final salvo: {JSON_FINAL.name}")
    log_cb("Processo concluído com sucesso!")