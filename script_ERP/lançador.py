import os
import re
import time
import json
import pyodbc
import threading
import tkinter as tk
from tkinter import filedialog, ttk
import pyautogui
import xml.etree.ElementTree as ET
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import cv2
import numpy as np
from PIL import ImageGrab, ImageTk

# ─── Paleta de cores ──────────────────────────────────────────────────────────
BG        = "#0f1117"
PANEL     = "#1a1d27"
ACCENT    = "#4f8ef7"
SUCCESS   = "#2ecc71"
DANGER    = "#e74c3c"
WARNING   = "#f39c12"
TEXT      = "#e8eaf0"
SUBTEXT   = "#8b93a8"
BORDER    = "#2a2d3e"
BTN_BG    = "#252836"
BTN_HOV   = "#2f3347"
ENTRY_BG  = "#12151f"
GREEN_BTN = "#1a6e3c"
GREEN_HOV = "#1f8a4a"


# ─── Utilitário: lê um XML de NF-e e devolve metadados essenciais ─────────────
def parse_nfe_xml(caminho_xml: str) -> dict:
    """Extrai do XML apenas o necessário: CNPJ, Chave, Valor, Emissão e contagem de produtos."""
    dados = {
        "fornecedor_cnpj": "", "chave": "", "valor_nota": "", 
        "numero_nota": "", "fornecedor_nome": "", "data_emissao": "", 
        "inf_cpl": "", "inf_ad_fisco": "", "produtos": []
    }
    
    if not caminho_xml or not os.path.exists(caminho_xml):
        return dados

    NS = "http://www.portalfiscal.inf.br/nfe"

    try:
        tree = ET.parse(caminho_xml)
        root = tree.getroot()

        # 1. CNPJ do Emitente
        cnpj_node = root.find(f".//{{{NS}}}emit/{{{NS}}}CNPJ")
        if cnpj_node is not None and cnpj_node.text:
            dados["fornecedor_cnpj"] = cnpj_node.text

        # 2. Chave da NF-e
        chave_node = root.find(f".//{{{NS}}}protNFe/{{{NS}}}infProt/{{{NS}}}chNFe")
        if chave_node is not None and chave_node.text:
            dados["chave"] = chave_node.text
        else:
            infNFe = root.find(f".//{{{NS}}}infNFe")
            if infNFe is not None and infNFe.get("Id"):
                dados["chave"] = infNFe.get("Id").replace("NFe", "")

        # 3. Valor da Nota (Tenta o vPag, senão pega vNF)
        vpag_node = root.find(f".//{{{NS}}}pag/{{{NS}}}detPag/{{{NS}}}vNF")
        if vpag_node is not None and vpag_node.text:
            dados["valor_nota"] = vpag_node.text
        else:
            vnf_node = root.find(f".//{{{NS}}}total/{{{NS}}}ICMSTot/{{{NS}}}vNF")
            if vnf_node is not None and vnf_node.text:
                dados["valor_nota"] = vnf_node.text

        # 4. Dados básicos pro log
        ide = root.find(f".//{{{NS}}}ide")
        if ide is not None:
            nnf_node = ide.find(f"{{{NS}}}nNF")
            if nnf_node is not None and nnf_node.text: 
                dados["numero_nota"] = nnf_node.text
            dhemi = ide.find(f"{{{NS}}}dhEmi")
            if dhemi is not None and dhemi.text: 
                dados["data_emissao"] = dhemi.text[:10]

        emit = root.find(f".//{{{NS}}}emit")
        if emit is not None:
            xnome = emit.find(f"{{{NS}}}xNome")
            if xnome is not None and xnome.text: 
                dados["fornecedor_nome"] = xnome.text

        # 5. Complementos (necessário para a digitação da macro)
        infAdic = root.find(f".//{{{NS}}}infAdic")
        if infAdic is not None:
            infCpl = infAdic.find(f"{{{NS}}}infCpl")
            if infCpl is not None and infCpl.text: 
                dados["inf_cpl"] = infCpl.text
            infAdFisco = infAdic.find(f"{{{NS}}}infAdFisco")
            if infAdFisco is not None and infAdFisco.text: 
                dados["inf_ad_fisco"] = infAdFisco.text

        # 6. Contagem de produtos para o robô saber quantas bolinhas ler
        for det in root.findall(f".//{{{NS}}}det"):
            dados["produtos"].append({})

    except Exception as e:
        print(f"Erro ao processar o XML {caminho_xml}: {e}")

    return dados


def calcular_vencimento_mes_seguinte(data_compra: str) -> str:
    """
    Vencimento = dia 20 do mês seguinte a 'data_compra'.
    Aceita 'data_compra' em 'AAAA-MM-DD' ou 'DD/MM/AAAA'.
    """
    if not data_compra:
        return ""

    dt = None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            dt = datetime.strptime(data_compra, fmt)
            break
        except ValueError:
            continue

    if dt is None:
        return ""

    venc = (dt + relativedelta(months=1)).replace(day=20)
    return venc.strftime("%d/%m/%Y")


def _to_float(valor) -> "float | None":
    if valor is None or valor == "":
        return None
    try:
        return float(str(valor).replace(",", "."))
    except (ValueError, TypeError):
        return None


# ─── Detecção de bolinhas via OpenCV/HSV ──────────────────────────────────────
_VERDE_MIN  = np.array([ 40,  50,  50])
_VERDE_MAX  = np.array([ 85, 255, 255])
_VERM_MIN1  = np.array([  0,  50,  50])
_VERM_MAX1  = np.array([ 10, 255, 255])
_VERM_MIN2  = np.array([170,  50,  50])
_VERM_MAX2  = np.array([180, 255, 255])
_AZUL_MIN   = np.array([100,  50,  50])
_AZUL_MAX   = np.array([130, 255, 255])
_AMAR_MIN   = np.array([ 25, 230,  50])
_AMAR_MAX   = np.array([ 40, 157, 255])


def conectar_banco():
    conn_str = (""
    )
    return pyodbc.connect(conn_str, timeout=5)


def fornecedor_existe(cnpj: str):
    cnpj_limpo = re.sub(r"\D", "", str(cnpj or ""))
    if not cnpj_limpo: return None
    try:
        conn = conectar_banco()
        cursor = conn.cursor()
        query = """
            SELECT TOP 1 CodCliente, CGCCPF
            FROM dbo.Clientes
            WHERE REPLACE(REPLACE(REPLACE(ISNULL(CGCCPF, ''), '.', ''), '/', ''), '-', '') = ?
        """
        cursor.execute(query, [cnpj_limpo])
        resultado = cursor.fetchone()
        cursor.close()
        conn.close()
        return str(resultado[0]) if resultado else None
    except Exception as e:
        print(f"Erro ao consultar fornecedor no banco: {e}")
        return None


def nota_ja_lancada(chave_nfe: str):
    chave_limpa = re.sub(r"\D", "", str(chave_nfe or ""))
    if not chave_limpa: return False
    try:
        conn = conectar_banco()
        cursor = conn.cursor()
        query = "SELECT TOP 1 1 FROM dbo.NotasFiscaisEntrada WHERE ChaveNFe = ?"
        cursor.execute(query, [chave_limpa])
        resultado = cursor.fetchone()
        cursor.close()
        conn.close()
        return resultado is not None
    except Exception as e:
        print(f"Erro ao consultar NotasFiscaisEntrada no banco: {e}")
        return False


def checar_fornecedor_e_nota(cnpj: str, chave_nfe: str) -> dict:
    cod_cliente = fornecedor_existe(cnpj)
    fornecedor_encontrado = cod_cliente is not None

    nota_lancada = False
    if fornecedor_encontrado or chave_nfe:
        nota_lancada = nota_ja_lancada(chave_nfe)

    return {
        "fornecedor_encontrado": fornecedor_encontrado,
        "cod_cliente":           cod_cliente,
        "nota_lancada":          nota_lancada,
    }


def capturar_hsv_area(x1: int, y1: int, x2: int, y2: int):
    shot = ImageGrab.grab(bbox=(x1, y1, x2, y2))
    bgr  = cv2.cvtColor(np.array(shot), cv2.COLOR_RGB2BGR)
    return cv2.cvtColor(bgr, cv2.COLOR_BGR2HSV)


def _maior_contorno_valido(mask, min_size: int = 5):
    cnts, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    for c in cnts:
        x, y, w, h = cv2.boundingRect(c)
        if w >= min_size and h >= min_size: return True
    return False


def classificar_bolinha_hsv(x1: int, y1: int, x2: int, y2: int) -> str:
    hsv = capturar_hsv_area(x1, y1, x2, y2)
    mask_verm  = cv2.inRange(hsv, _VERM_MIN1, _VERM_MAX1) | cv2.inRange(hsv, _VERM_MIN2, _VERM_MAX2)
    mask_amar  = cv2.inRange(hsv, _AMAR_MIN,  _AMAR_MAX)
    mask_verde = cv2.inRange(hsv, _VERDE_MIN, _VERDE_MAX)
    mask_azul  = cv2.inRange(hsv, _AZUL_MIN,  _AZUL_MAX)

    if _maior_contorno_valido(mask_verm): return "vermelho"
    if _maior_contorno_valido(mask_amar): return "amarelo"
    if _maior_contorno_valido(mask_verde): return "verde"
    if _maior_contorno_valido(mask_azul): return "azul"
    return "desconhecido"


class NFeXMLMacro:
    def __init__(self, root):
        self.root = root
        self.root.title("Pluton — Importação de NF-e (genérico)")
        self.root.configure(bg=BG)
        self.root.resizable(True, True)
        self.root.minsize(900, 740)

        self.arquivo_json   = ""   # Caminho do arquivo JSON principal
        self.inputs_manuais = {}   
        self.posicoes       = {}
        self.usar_padrao    = tk.BooleanVar(value=True)
        self.pausado        = False
        self.resultados     = []
        self._stats         = {"total": 0, "ok": 0, "fail": 0}

        self.delay_apos_xml    = tk.StringVar(value="2.0")
        self.delay_cadastro    = tk.StringVar(value="3.5")
        self.delay_entre_notas = tk.StringVar(value="2.0")

        self._apply_style()
        self._build_ui()
        self.root.bind("<F8>", self.toggle_pausa)

    def _apply_style(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TFrame",        background=BG)
        style.configure("Panel.TFrame",  background=PANEL)
        style.configure("TLabel",        background=BG,    foreground=TEXT,    font=("Segoe UI", 10))
        style.configure("Sub.TLabel",    background=PANEL, foreground=SUBTEXT, font=("Segoe UI", 9))
        style.configure("Head.TLabel",   background=BG,    foreground=TEXT,    font=("Segoe UI", 13, "bold"))
        style.configure("Accent.TLabel", background=BG,    foreground=ACCENT,  font=("Segoe UI", 10, "bold"))
        style.configure("TCheckbutton",  background=PANEL, foreground=TEXT,    font=("Segoe UI", 10), indicatorcolor=ACCENT)
        style.map("TCheckbutton",        background=[("active", PANEL)])
        style.configure("TEntry",        fieldbackground=ENTRY_BG, foreground=TEXT, insertcolor=TEXT,
                        bordercolor=BORDER, relief="flat", font=("Segoe UI", 10))
        style.map("TEntry",              bordercolor=[("focus", ACCENT)])
        style.configure("Vertical.TScrollbar", background=PANEL, troughcolor=BG, arrowcolor=SUBTEXT, bordercolor=BG)

    def _build_ui(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)
        self._build_header()

        pane = ttk.Frame(self.root)
        pane.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 16))
        pane.columnconfigure(0, weight=0, minsize=310)
        pane.columnconfigure(1, weight=1)
        pane.rowconfigure(0, weight=1)

        self._build_left(pane)
        self._build_right(pane)
        self._build_status_bar()

    def _build_header(self):
        hdr = tk.Frame(self.root, bg=PANEL, height=56)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.grid_propagate(False)

        canvas = tk.Canvas(hdr, bg=PANEL, highlightthickness=0, height=4)
        canvas.place(x=0, y=52, relwidth=1)
        canvas.create_rectangle(0, 0, 10000, 4, fill=ACCENT, outline="")

        tk.Label(hdr, text="⚡  Pluton", bg=PANEL, fg=ACCENT, font=("Segoe UI", 16, "bold")).pack(side="left", padx=20, pady=10)
        tk.Label(hdr, text="Importação de NF-e — Genérico", bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 10)).pack(side="left", pady=10)

        self.lbl_pause = tk.Label(hdr, text="▶  RODANDO", bg=PANEL, fg=SUCCESS, font=("Segoe UI", 9, "bold"))
        self.lbl_pause.pack(side="right", padx=20)
        tk.Label(hdr, text="F8 para pausar / retomar   |", bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).pack(side="right")

    def _build_left(self, parent):
        wrapper = tk.Frame(parent, bg=BG)
        wrapper.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        wrapper.columnconfigure(0, weight=1)
        wrapper.rowconfigure(0, weight=1)

        canvas = tk.Canvas(wrapper, bg=BG, highlightthickness=0, bd=0)
        canvas.grid(row=0, column=0, sticky="nsew")

        vsb = ttk.Scrollbar(wrapper, orient="vertical", command=canvas.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        canvas.configure(yscrollcommand=vsb.set)

        left = ttk.Frame(canvas)
        left.columnconfigure(0, weight=1)
        left_win = canvas.create_window((0, 0), window=left, anchor="nw")

        def _on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_configure(event):
            canvas.itemconfig(left_win, width=event.width)

        left.bind("<Configure>", _on_frame_configure)
        canvas.bind("<Configure>", _on_canvas_configure)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll( 1, "units"))

        # ── Arquivo de Pedidos (JSON) ──────────────────────────────────────────────
        self._section(left, "🧾  Arquivo de Pedidos (JSON)", 0)
        self.lbl_arquivo_json = self._info_row(left, "Arquivo JSON principal:", "Nenhum selecionado", 1)
        self._btn(left, "Selecionar Arquivo .json", self.selecionar_arquivo_json, 2)

        ttk.Separator(left, orient="horizontal").grid(row=3, column=0, sticky="ew", pady=10)

        # ── Posições de clique ───────────────────────────────────────────────
        self._section(left, "🖱  Posições de clique", 4)

        chk = ttk.Checkbutton(left, text="Usar coordenadas padrão", variable=self.usar_padrao)
        chk.grid(row=5, column=0, sticky="w", pady=(0, 6))

        campos = [
            ("importar_xml",    "Botão Importar XML (tela 1)"),
            ("tres_pontos",     "Botão ... (tela 2)"),
            ("btn_adicionar",   "Botão add produto — ícone verde (tela 3)"),
            ("btn_gravar",      "Botão Gravar (cadastro de produto)"),
            ("lista_topo_esq",  "Lista produtos: canto SUPERIOR ESQ"),
            ("lista_topo_dir",  "Lista produtos: canto SUPERIOR DIR"),
            ("btn_scroll_down", "Botão scroll ↓ da lista"),
            ("confirmar",       "Botão Confirmar (tela 3)"),
            ("confirmar2",      "Botão Confirmar (tela 4)"),
            ("anexar2",         "Botão anexar — ícone clipe"),
            ("anexar",          "Campo de caminho do anexo"),
            ("pgto",            "Botão de pagamento (1º clique)"),
            ("pgto2",           "Botão de pagamento (2º clique)"),
        ]

        container = ttk.Frame(left, style="Panel.TFrame", padding=10)
        container.grid(row=6, column=0, sticky="ew")
        container.columnconfigure(1, weight=1)
        container.columnconfigure(3, weight=1)

        for i, (campo, legenda) in enumerate(campos):
            ttk.Label(container, text=legenda, style="Sub.TLabel",
                      width=28, anchor="e").grid(row=i, column=0, padx=(0, 6), pady=3, sticky="e")
            x_var, y_var = tk.StringVar(), tk.StringVar()
            self.inputs_manuais[campo] = (x_var, y_var)
            ttk.Label(container, text="X", style="Sub.TLabel").grid(row=i, column=1, sticky="e")
            ttk.Entry(container, textvariable=x_var, width=6).grid(row=i, column=2, padx=3)
            ttk.Label(container, text="Y", style="Sub.TLabel").grid(row=i, column=3, sticky="e")
            ttk.Entry(container, textvariable=y_var, width=6).grid(row=i, column=4, padx=(3, 8))
            btn = tk.Button(container, text="📍", bg=BTN_BG, fg=ACCENT, relief="flat",
                            cursor="hand2", font=("Segoe UI", 10),
                            command=lambda c=campo: self.definir_posicao(c))
            btn.grid(row=i, column=5, pady=2)
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg=BTN_HOV))
            btn.bind("<Leave>", lambda e, b=btn: b.config(bg=BTN_BG))

        ttk.Separator(left, orient="horizontal").grid(row=7, column=0, sticky="ew", pady=10)

        # ── Configurações de tempo ─────────────────────────────────────────────
        self._section(left, "⏱  Tempos (segundos)", 8)
        delay_frame = ttk.Frame(left, style="Panel.TFrame", padding=10)
        delay_frame.grid(row=9, column=0, sticky="ew")
        delay_frame.columnconfigure(1, weight=1)

        for i, (label, var) in enumerate([
            ("Após carregar XML:", self.delay_apos_xml),
            ("Cadastro produto:", self.delay_cadastro),
            ("Entre notas:",      self.delay_entre_notas),
        ]):
            ttk.Label(delay_frame, text=label, style="Sub.TLabel").grid(row=i, column=0, sticky="w", pady=2)
            ttk.Entry(delay_frame, textvariable=var, width=6).grid(row=i, column=1, sticky="w", padx=8)

        ttk.Separator(left, orient="horizontal").grid(row=10, column=0, sticky="ew", pady=10)

        # ── Vencimento automático ──────────────────────────────────────────────
        self._section(left, "📅  Vencimento automático", 11)
        venc_frame = ttk.Frame(left, style="Panel.TFrame", padding=10)
        venc_frame.grid(row=12, column=0, sticky="ew")
        venc_frame.columnconfigure(0, weight=1)
        ttk.Label(venc_frame, style="Sub.TLabel", justify="left",
                  text="Vencimento = dia 20 do mês seguinte à data_compra do pedido.").grid(row=0, column=0, sticky="w")

        ttk.Separator(left, orient="horizontal").grid(row=13, column=0, sticky="ew", pady=10)

        # ── Botões de ação ─────────────────────────────────────────────────────
        self._btn(left, "🔍  Pré-visualizar XMLs (sem executar)",
                  lambda: threading.Thread(target=self.previsualizar, daemon=True).start(),
                  14, color=GREEN_BTN, text_color="white")

        self._btn(left, "🎯  Preview da área de bolinhas",
                  self.mostrar_preview_bolinhas,
                  15, color=BTN_BG, text_color=ACCENT)

        self._btn(left, "▶  Executar importação em massa",
                  lambda: threading.Thread(target=self.executar_macro, daemon=True).start(),
                  16, color=ACCENT, text_color="white", big=True)

        # ── Estatísticas ───────────────────────────────────────────────────────
        stats_frame = ttk.Frame(left, style="Panel.TFrame", padding=10)
        stats_frame.grid(row=17, column=0, sticky="ew", pady=(10, 0))
        stats_frame.columnconfigure((0, 1, 2), weight=1)

        for col, (label, color, attr) in enumerate([
            ("Total",   SUBTEXT, "lbl_total"),
            ("✅ OK",   SUCCESS, "lbl_ok"),
            ("❌ Falha", DANGER, "lbl_fail"),
        ]):
            f = tk.Frame(stats_frame, bg=ENTRY_BG, padx=10, pady=8)
            f.grid(row=0, column=col, padx=4, sticky="ew")
            tk.Label(f, text=label, bg=ENTRY_BG, fg=SUBTEXT, font=("Segoe UI", 8)).pack()
            lbl = tk.Label(f, text="0", bg=ENTRY_BG, fg=color, font=("Segoe UI", 18, "bold"))
            lbl.pack()
            setattr(self, attr, lbl)

    def _build_right(self, parent):
        right = ttk.Frame(parent)
        right.grid(row=0, column=1, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)

        self._section(right, "📋  Log de execução", 0)

        log_frame = tk.Frame(right, bg=ENTRY_BG, relief="flat", bd=0)
        log_frame.grid(row=1, column=0, sticky="nsew")
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        self.output = tk.Text(log_frame, bg=ENTRY_BG, fg=TEXT, insertbackground=TEXT,
                              font=("Cascadia Code", 9), relief="flat", bd=0,
                              padx=12, pady=10, wrap="word")
        self.output.grid(row=0, column=0, sticky="nsew")

        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.output.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.output.configure(yscrollcommand=scrollbar.set)

        self.output.tag_configure("ok",    foreground=SUCCESS)
        self.output.tag_configure("error", foreground=DANGER)
        self.output.tag_configure("warn",  foreground=WARNING)
        self.output.tag_configure("info",  foreground=ACCENT)
        self.output.tag_configure("sub",   foreground=SUBTEXT)
        self.output.tag_configure("green", foreground="#27ae60")
        self.output.tag_configure("bold",  foreground=TEXT, font=("Cascadia Code", 9, "bold"))

        btn_frame = ttk.Frame(right)
        btn_frame.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        self._btn(btn_frame, "🗑  Limpar log", self._clear_log, 0, side="left", small=True)

    def _build_status_bar(self):
        bar = tk.Frame(self.root, bg=PANEL, height=26)
        bar.grid(row=2, column=0, sticky="ew")
        self.status_var = tk.StringVar(value="Pronto")
        tk.Label(bar, textvariable=self.status_var, bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 8), anchor="w").pack(side="left", padx=12, pady=4)

    # ── Helpers UI ─────────────────────────────────────────────────────────────
    def _section(self, parent, title, row):
        ttk.Label(parent, text=title, style="Head.TLabel").grid(row=row, column=0, sticky="w", pady=(10, 4))

    def _info_row(self, parent, label, default, row):
        f = tk.Frame(parent, bg=ENTRY_BG, padx=10, pady=6)
        f.grid(row=row, column=0, sticky="ew", pady=2)
        tk.Label(f, text=label, bg=ENTRY_BG, fg=SUBTEXT, font=("Segoe UI", 8)).pack(anchor="w")
        lbl = tk.Label(f, text=default, bg=ENTRY_BG, fg=TEXT, font=("Segoe UI", 9),
                       wraplength=260, anchor="w", justify="left")
        lbl.pack(anchor="w")
        return lbl

    def _btn(self, parent, text, cmd, row=None, color=BTN_BG, text_color=TEXT,
             big=False, side=None, small=False):
        font_size = 11 if big else (9 if small else 10)
        weight    = "bold" if big else "normal"
        pady_val  = 10 if big else (4 if small else 7)
        btn = tk.Button(parent, text=text, command=cmd, bg=color, fg=text_color,
                        relief="flat", cursor="hand2", font=("Segoe UI", font_size, weight),
                        activebackground=BTN_HOV, activeforeground=text_color,
                        pady=pady_val, padx=12, bd=0)
        hover_bg = BTN_HOV if color == BTN_BG else color
        btn.bind("<Enter>", lambda e: btn.config(bg=hover_bg))
        btn.bind("<Leave>", lambda e: btn.config(bg=color))
        if side:
            btn.pack(side=side, padx=(0, 8))
        else:
            btn.grid(row=row, column=0, sticky="ew", pady=(2, 0))
        return btn

    def _log(self, msg, tag=""):
        self.output.insert(tk.END, msg + "\n", tag)
        self.output.see(tk.END)

    def _clear_log(self):
        self.output.delete("1.0", tk.END)

    def _set_status(self, msg):
        self.status_var.set(msg)
        self.root.update_idletasks()

    def _update_stats(self):
        self.lbl_total.config(text=str(self._stats["total"]))
        self.lbl_ok.config(text=str(self._stats["ok"]))
        self.lbl_fail.config(text=str(self._stats["fail"]))

    def _wait_pause(self):
        while self.pausado:
            time.sleep(0.3)

    def _delay(self, var: tk.StringVar) -> float:
        try:
            return float(var.get())
        except ValueError:
            return 2.0

    # ── Seleção de Arquivo JSON ─────────────────────────────────────────────────
    def selecionar_arquivo_json(self):
        path = filedialog.askopenfilename(filetypes=[("JSON Files", "*.json")])
        if path:
            self.arquivo_json = path
            self.lbl_arquivo_json.config(text=os.path.basename(path) or path)
            self._log(f"🧾 Arquivo JSON selecionado: {path}", "info")


    # ── Captura de posição ──────────────────────────────────────────────────────
    def definir_posicao(self, campo):
        self.root.withdraw()
        pyautogui.confirm(
            text=f'Posicione o mouse sobre "{campo}" e clique em OK.',
            title='Capturar posição', buttons=['OK'])
        pos = pyautogui.position()
        self.posicoes[campo] = pos
        self.inputs_manuais[campo][0].set(str(pos.x))
        self.inputs_manuais[campo][1].set(str(pos.y))
        self.root.deiconify()

    def get_posicao(self, campo) -> tuple:
        padroes = {
            "importar_xml":    (1205, 178),
            "tres_pontos":     (1126, 183),
            "btn_adicionar":   (1380, 549),
            "btn_gravar":      (1118, 673),
            "lista_topo_esq":  (530, 377),
            "lista_topo_dir":  (551, 517),
            "btn_scroll_down": (1382, 509),
            "confirmar":       (1325, 771),
            "confirmar2":      (1139, 824),
            "anexar2":         (722, 148),
            "anexar":          (932, 228),
            "pgto":            (836, 598),
            "pgto2":           (849, 676),
        }
        if self.usar_padrao.get():
            return padroes.get(campo, (0, 0))
        x_var, y_var = self.inputs_manuais[campo]
        try:
            return (int(x_var.get()), int(y_var.get()))
        except ValueError:
            return self.posicoes.get(campo, padroes.get(campo, (0, 0)))

    def mostrar_preview_bolinhas(self):
        try:
            x1, y1 = self.get_posicao("lista_topo_esq")
            x2, _  = self.get_posicao("lista_topo_dir")
            y2     = y1 + 8 * 17
        except Exception as e:
            self._log(f"❌ Erro ao capturar preview: {str(e)}", "error")
            return
        shot   = ImageGrab.grab(bbox=(x1, y1, x2, y2))
        win    = tk.Toplevel(self.root)
        win.title("Preview — área de bolinhas")
        win.configure(bg=BG)
        img_tk = ImageTk.PhotoImage(shot)
        lbl    = tk.Label(win, image=img_tk, bg=BG)
        lbl.image = img_tk
        lbl.pack(padx=10, pady=10)
        tk.Label(win, text=f"Região: ({x1},{y1}) → ({x2},{y2})",
                 bg=BG, fg=SUBTEXT, font=("Segoe UI", 8)).pack(pady=(0, 8))

    def toggle_pausa(self, event=None):
        self.pausado = not self.pausado
        if self.pausado:
            self.lbl_pause.config(text="⏸  PAUSADO", fg=WARNING)
            self._log("⏸️  Macro pausado — F8 para continuar.", "warn")
        else:
            self.lbl_pause.config(text="▶  RODANDO", fg=SUCCESS)
            self._log("▶️  Macro retomado.", "ok")

    # ═══════════════════════════════════════════════════════════════════════════
    #  PRÉ-VISUALIZAÇÃO (via arquivo JSON)
    # ═══════════════════════════════════════════════════════════════════════════
    def previsualizar(self):
        if not getattr(self, 'arquivo_json', None):
            self._log("❌ Nenhum arquivo JSON selecionado!", "error"); return

        try:
            with open(self.arquivo_json, "r", encoding="utf-8") as f:
                conteudo = json.load(f)
            pedidos = conteudo if isinstance(conteudo, list) else [conteudo]
        except Exception as e:
            self._log(f"❌ Erro ao ler JSON: {str(e)}", "error"); return

        self._log(f"\n{'═'*52}", "sub")
        self._log(f"🔍  PRÉ-VISUALIZAÇÃO (via JSON) — {len(pedidos)} pedido(s)", "green")
        self._log(f"{'═'*52}", "sub")

        resultados_preview = []

        for i, pedido in enumerate(pedidos, 1):
            if not isinstance(pedido, dict): continue
                
            pedido_id        = pedido.get("pedido_id", f"Item_{i}")
            id_pagamento     = pedido.get("id_pagamento", "")  # <--- NOVO CAMPO AQUI
            data_compra      = pedido.get("data_compra", "")
            caminho_xml      = pedido.get("caminho_xml", "")
            nome_xml         = os.path.basename(caminho_xml) if caminho_xml else "Sem XML"
            cartao           = pedido.get("cartao_utilizado", "")
            parcelas         = pedido.get("parcelas", "")
            conta_resultados = pedido.get("contaResultados", "")
            
            # Formata o ajuste financeiro
            ajuste   = pedido.get("ajuste_financeiro", {})
            aj_tipo  = ajuste.get("tipo", "nenhum")
            aj_valor = _to_float(ajuste.get("valor", 0)) or 0.0
            aj_fmt   = f"{aj_tipo} (R$ {aj_valor:.2f})" if aj_valor > 0 else "nenhum (R$ 0.00)"

            self._log(f"\n🧾 [{i}/{len(pedidos)}] Pedido: {pedido_id} | PGTO: {id_pagamento} | Compra: {data_compra}", "bold")

            linha = {
                "Pedido ID":          pedido_id,
                "ID Pagamento":       id_pagamento,  # <--- NOVA COLUNA AQUI
                "Data Compra":        data_compra,
                "XML":                nome_xml,
                "Cartão":             cartao,
                "Parcelas":           parcelas,
                "Conta Resultados":   conta_resultados,
                "Fornecedor":         "",
                "CNPJ":               "",
                "Chave NF-e":         "",
                "Valor Nota (XML)":   "",
                "Valor Pedido (JSON)": f"{_to_float(pedido.get('valor_total', 0)):.2f}",
                "Ajuste Financeiro":  aj_fmt,
                "Diferença":          "",
                "Fornecedor OK?":     "",
                "Nota Lançada?":      "",
                "Status Geral":       "",
                "Observação":         "",
            }

            if not caminho_xml or not os.path.exists(caminho_xml):
                self._log(f"   ❌ Arquivo XML não encontrado no PC: {caminho_xml}", "error")
                linha.update({"Status Geral": "❌ XML não encontrado", "Observação": "Caminho_xml no JSON não existe no PC."})
                resultados_preview.append(linha)
                continue

            try:
                dados = parse_nfe_xml(caminho_xml)
                
                cnpj  = dados.get("fornecedor_cnpj", "")
                chave = dados.get("chave", "")

                self._log(f"   Fornecedor : {dados.get('fornecedor_nome', 'N/A')} (CNPJ {cnpj})", "sub")
                self._log(f"   Chave NF-e : {chave}", "sub")

                valor_json = _to_float(pedido.get("valor_total", 0))
                valor_nfe  = _to_float(dados.get("valor_nota", "0"))
                obs = ""
                diferenca = ""

                if valor_json is not None and valor_nfe is not None:
                    diff_val = round(valor_nfe - valor_json, 2)
                    diferenca = str(diff_val)
                    if abs(diff_val) > 0.01:
                        diferenca = "⚠️ " + diferenca
                        obs = f"JSON (R${valor_json:.2f}) / XML (R${valor_nfe:.2f})"

                checks = checar_fornecedor_e_nota(cnpj, chave)
                fornecedor_ok = checks["fornecedor_encontrado"]
                nota_lancada  = checks["nota_lancada"]

                if fornecedor_ok:
                    self._log(f"   ✅ Fornecedor no BD (Cod={checks['cod_cliente']}).", "ok")
                else:
                    self._log(f"   ❌ Fornecedor NÃO cadastrado no BD.", "error")

                if nota_lancada:
                    self._log(f"   ⏭️  Nota já lançada.", "warn")
                else:
                    self._log(f"   ⬜ Nota ainda não lançada.", "sub")

                if not fornecedor_ok:
                    status_geral = "❌ Fornecedor não cadastrado"
                elif nota_lancada:
                    status_geral = "⏭️ Já lançada"
                else:
                    status_geral = "🟢 Pronta para lançar"

                linha.update({
                    "Fornecedor":           dados.get("fornecedor_nome", ""),
                    "CNPJ":                 cnpj,
                    "Chave NF-e":           chave,
                    "Valor Nota (XML)":     dados.get("valor_nota", ""),
                    "Diferença":            diferenca,
                    "Fornecedor OK?":       "Sim" if fornecedor_ok else "Não",
                    "Nota Lançada?":        "Sim" if nota_lancada else "Não",
                    "Status Geral":         status_geral,
                    "Observação":           obs,
                })

            except Exception as e:
                self._log(f"   ❌ Erro na leitura do XML: {str(e)}", "error")
                linha.update({"Status Geral": "❌ Erro interno", "Observação": str(e)})

            resultados_preview.append(linha)

        self._log(f"\n{'═'*52}", "sub")
        self._log("✅  Pré-visualização concluída.", "ok")
        self._set_status("Pré-visualização concluída")

        self._gerar_relatorio(resultados_preview, aba="Preview via JSON", prefixo_arquivo="preview_via_json")

    # ═══════════════════════════════════════════════════════════════════════════
    #  MACRO PRINCIPAL (via arquivo JSON)
    # ═══════════════════════════════════════════════════════════════════════════

    def executar_macro(self):
        if not getattr(self, 'arquivo_json', None):
            self._log("❌ Nenhum arquivo JSON selecionado!", "error"); return

        try:
            with open(self.arquivo_json, "r", encoding="utf-8") as f:
                conteudo = json.load(f)
            pedidos = conteudo if isinstance(conteudo, list) else [conteudo]
        except Exception as e:
            self._log(f"❌ Erro ao ler JSON: {str(e)}", "error"); return

        self._stats = {"total": len(pedidos), "ok": 0, "fail": 0}
        self._update_stats()
        self.resultados = []

        self._log(f"\n{'═'*52}", "sub")
        self._log(f"🚀  INICIANDO (via JSON) — {len(pedidos)} pedido(s)", "green")
        self._log(f"{'═'*52}\n", "sub")
        self._log("⚠️  Coloque o foco na janela do sistema antes de iniciar.", "warn")
        time.sleep(3)

        for i, pedido in enumerate(pedidos, 1):
            if not isinstance(pedido, dict): continue

            self._wait_pause()
            pedido_id        = pedido.get("pedido_id", f"Item_{i}")
            id_pagamento     = pedido.get("id_pagamento", "")  # <--- NOVO CAMPO AQUI
            data_compra      = pedido.get("data_compra", "")
            caminho_xml      = pedido.get("caminho_xml", "")
            nome_xml         = os.path.basename(caminho_xml) if caminho_xml else "Sem XML"
            cartao           = pedido.get("cartao_utilizado", "")
            parcelas         = pedido.get("parcelas", "")
            conta_resultados = pedido.get("contaResultados", "")
            
            ajuste   = pedido.get("ajuste_financeiro", {})
            aj_tipo  = ajuste.get("tipo", "nenhum")
            aj_valor = _to_float(ajuste.get("valor", 0)) or 0.0
            aj_fmt   = f"{aj_tipo} (R$ {aj_valor:.2f})" if aj_valor > 0 else "nenhum (R$ 0.00)"

            self._log(f"\n🧾 [{i}/{len(pedidos)}] Pedido: {pedido_id} | XML: {nome_xml}", "bold")
            self._set_status(f"Processando {pedido_id} ({i}/{len(pedidos)})")

            # Monta a base da linha para o relatório
            linha_res = {
                "Pedido": pedido_id, 
                "ID Pagamento": id_pagamento,  # <--- NOVA COLUNA AQUI
                "Data Compra": data_compra,
                "XML": nome_xml, 
                "Cartão": cartao,
                "Parcelas": parcelas,
                "Conta Resultados": conta_resultados,
                "Valor Pedido (JSON)": f"{_to_float(pedido.get('valor_total', 0)):.2f}",
                "Ajuste Financeiro": aj_fmt
            }

            if not caminho_xml or not os.path.exists(caminho_xml):
                self._log(f"   ❌ XML não encontrado: {caminho_xml}", "error")
                self._stats["fail"] += 1
                linha_res.update({"Status": "❌ Falha", "Motivo": "XML não encontrado pelo caminho no JSON"})
                self.resultados.append(linha_res)
                self._update_stats()
                continue

            try:
                dados = parse_nfe_xml(caminho_xml)
                dados["vencimento"] = calcular_vencimento_mes_seguinte(data_compra or dados.get("data_emissao", ""))

                status, motivo = self._importar_nota(caminho_xml, dados, nome_xml)

                linha_res.update({
                    "Nota": dados.get("numero_nota", ""),
                    "Fornecedor": dados.get("fornecedor_nome", ""),
                    "Chave": dados.get("chave", ""), 
                    "Vencimento": dados.get("vencimento", ""), 
                    "Valor Nota (XML)": dados.get("valor_nota", "")
                })

                if status == "ok":
                    self._stats["ok"] += 1
                    linha_res.update({"Status": "✅ OK", "Motivo": ""})
                elif status == "pulado":
                    self._stats["ok"] += 1
                    linha_res.update({"Status": "⏭️ Pulado", "Motivo": motivo})
                else:
                    self._stats["fail"] += 1
                    linha_res.update({"Status": "❌ Falha", "Motivo": motivo})
                
                self.resultados.append(linha_res)

            except Exception as e:
                self._log(f"   ❌ Erro durante o fluxo da nota: {str(e)}", "error")
                self._stats["fail"] += 1
                linha_res.update({"Status": "❌ Falha", "Motivo": str(e)})
                self.resultados.append(linha_res)

            self._update_stats()
            time.sleep(self._delay(self.delay_entre_notas))

        self._log(f"\n{'─'*50}", "sub")
        self._log(f"🏁 Concluído: {self._stats['ok']} OK  |  {self._stats['fail']} falha(s)", "bold")
        self._set_status("Concluído")
        self._gerar_relatorio(prefixo_arquivo="relatorio_nfe_via_json")

    # ── Validações de banco (CHECK 1 + CHECK 2) ─────────────────────────────────
    def _validar_lancamento(self, dados: dict):
        cnpj  = dados.get("fornecedor_cnpj", "")
        chave = dados.get("chave", "")

        cod_cliente = fornecedor_existe(cnpj)
        if cod_cliente is None:
            self._log(f"   ❌ Fornecedor não existe no sistema (CNPJ {cnpj}).", "error")
            return False, "fornecedor não existe no sistema"

        self._log(f"   ✅ Fornecedor encontrado (CodCliente={cod_cliente}, CNPJ={cnpj}).", "sub")

        if nota_ja_lancada(chave):
            self._log(f"   ⏭️  Nota {dados.get('numero_nota')} (chave {chave}) já lançada. Pulando.", "warn")
            return False, "nota já lançada"

        self._log(f"   ✅ Nota {dados.get('numero_nota')} ainda não lançada — seguindo.", "sub")
        return True, ""


    # ── Fluxo de uma nota ───────────────────────────────────────────────────────
    def _importar_nota(self, caminho_xml: str, dados: dict, nome_xml: str):
        delay_xml      = self._delay(self.delay_apos_xml)
        delay_cadastro = self._delay(self.delay_cadastro)

        pode_lancar, motivo = self._validar_lancamento(dados)
        if not pode_lancar:
            return "pulado", motivo

        self._log("   → Clicando em Importar XML...", "sub")
        pyautogui.click(self.get_posicao("importar_xml"))
        time.sleep(1.5)

        self._log("   → Clicando em '...' para selecionar arquivo...", "sub")
        pyautogui.click(self.get_posicao("tres_pontos"))
        time.sleep(1.0)

        nome_arquivo = os.path.basename(caminho_xml)
        pyautogui.typewrite(nome_arquivo, interval=0.03)
        pyautogui.press("enter")
        pyautogui.sleep(0.5)
        pyautogui.click(1316, 746)
        self._log(f"   → Arquivo digitado: {nome_arquivo}", "sub")

        time.sleep(delay_xml)

        n_produtos = len(dados["produtos"])
        self._log(f"   → Verificando {n_produtos} produto(s)...", "sub")
        self._verificar_e_cadastrar_produtos(n_produtos, delay_cadastro)

        self._log("   → Clicando em Confirmar (importação NF-e)...", "sub")
        pyautogui.click(self.get_posicao("confirmar"))
        time.sleep(1.0)
        pyautogui.press("enter")

        pendentes = getattr(self, "pendentes_confirmacao", 0)
        if pendentes:
            self._log(f"   → Confirmando {pendentes} caixa(s) de divergência...", "sub")
            for _ in range(pendentes):
                pyautogui.press("enter")
                time.sleep(0.5)

        time.sleep(1.0)

        self._log("   → Confirmando nota de entrada...", "sub")
        pyautogui.click(self.get_posicao("confirmar2"))
        pyautogui.click(self.get_posicao("confirmar2"))
        pyautogui.press("enter")
        time.sleep(0.5)

        texto_complemento = (dados.get("inf_cpl") or dados.get("inf_ad_fisco") or "").strip()
        if texto_complemento:
            self._log(f"   → Digitando complemento: {texto_complemento[:60]}{'...' if len(texto_complemento) > 60 else ''}", "sub")
            time.sleep(0.2)
            pyautogui.typewrite(texto_complemento, interval=0.03)
        else:
            self._log("   → Complemento (infCpl/infAdFisco) vazio, pulando...", "sub")

        self._log("   → Anexando XML...", "sub")
        pyautogui.click(self.get_posicao("anexar2"))
        time.sleep(0.5)
        pyautogui.click(self.get_posicao("anexar"))
        pyautogui.typewrite(caminho_xml, interval=0.03)
        pyautogui.press("enter")
        time.sleep(0.3)
        pyautogui.press("enter")

        self._log("   → Preenchendo pagamento...", "sub")
        pyautogui.click(self.get_posicao("pgto"))
        pyautogui.typewrite("Boleto")
        pyautogui.press("enter")
        pyautogui.press("enter")

        data_raw = dados.get("vencimento", "")
        if not data_raw:
            self._log("   ⚠️  Vencimento não calculado — pulando digitação da data.", "warn")
        else:
            data_digitavel = data_raw.replace("/", "")
            self._log(f"   → Digitando vencimento: {data_raw}", "sub")
            pyautogui.typewrite(data_digitavel, interval=0.05)

        for _ in range(4):
            pyautogui.press("enter")
        for _ in range(11):
            pyautogui.press("enter")
            time.sleep(0.1)

        time.sleep(5)
        pyautogui.press("enter")
        pyautogui.press("enter")
        pyautogui.press("enter")
        pyautogui.press("right")
        pyautogui.press("right")
        pyautogui.press("enter")
        time.sleep(1)

        self._log(f"   ✅ Nota {dados['numero_nota']} importada e lançada.", "ok")
        return "ok", ""


    # ── Verificação de bolinhas e cadastro de produtos ──────────────────────────
    def _verificar_e_cadastrar_produtos(self, n_produtos: int, delay_cadastro: float):
        LINHAS_VISIVEIS = 8
        MARGEM_TOPO     = 10
        ESPACO_LINHA    = 17

        x_esq, y_topo = self.get_posicao("lista_topo_esq")
        x_dir, _      = self.get_posicao("lista_topo_dir")
        largura_lista  = x_dir - x_esq

        y_primeira = y_topo + MARGEM_TOPO
        self._log(f"      Caixa: ({x_esq},{y_topo})→({x_dir},{y_topo+LINHAS_VISIVEIS*ESPACO_LINHA}) largura={largura_lista}px", "sub")

        self.pendentes_confirmacao = 0
        processados = 0

        while processados < n_produtos:
            restantes    = n_produtos - processados
            linhas_agora = min(LINHAS_VISIVEIS, restantes)

            for linha in range(linhas_agora):
                self._wait_pause()
                y_centro = y_primeira + linha * ESPACO_LINHA
                bx1 = x_esq
                by1 = y_centro - (ESPACO_LINHA // 2)
                bx2 = x_esq + 20
                by2 = y_centro + (ESPACO_LINHA // 2)
                time.sleep(0.15)

                cor = classificar_bolinha_hsv(bx1, by1, bx2, by2)
                idx = processados + linha + 1

                self._log(f"      Produto {idx:>2}: {cor.upper()}", "sub")

                if cor == "vermelho":
                    self._log(f"      ⚠️  Produto {idx} sem vínculo → abrindo cadastro...", "warn")
                    bx_centro = (bx1 + bx2) // 2
                    by_centro = (by1 + by2) // 2
                    pyautogui.moveTo(bx_centro, by_centro, duration=0.2)
                    pyautogui.click(bx_centro, by_centro)
                    time.sleep(0.3)
                    pyautogui.click(bx_centro, by_centro)
                    time.sleep(0.3)
                    pyautogui.click(bx_centro, by_centro)

                    pyautogui.click(self.get_posicao("btn_adicionar"))
                    time.sleep(delay_cadastro)

                    pyautogui.click(self.get_posicao("btn_gravar"))
                    time.sleep(5)

                    pyautogui.press("enter")
                    time.sleep(0.5)
                    pyautogui.press("enter")
                    time.sleep(1.5)
                    pyautogui.click(1316, 746)
                    self._log(f"      ✅ Produto {idx} cadastrado.", "ok")

                elif cor == "verde":
                    self._log(f"      ✅ Produto {idx} ok.", "ok")

                else:
                    bx_centro = (bx1 + bx2) // 2
                    by_centro = (by1 + by2) // 2
                    pyautogui.moveTo(bx_centro, by_centro, duration=0.2)
                    pyautogui.click(bx_centro, by_centro)
                    time.sleep(0.3)
                    pyautogui.click(bx_centro, by_centro)
                    time.sleep(0.3)
                    pyautogui.click(bx_centro, by_centro)
                    pyautogui.moveTo(941, 568, duration=0.2)
                    pyautogui.click(941, 568)
                    pyautogui.moveTo(980, 585, duration=0.2)
                    pyautogui.click(980, 585)
                    self._log(f"      ❓ Produto {idx} cor indeterminada — continuando.", "warn")

            processados += linhas_agora

            if processados < n_produtos:
                faltam = n_produtos - processados
                self._log(f"      ↓ Scroll — faltam {faltam} produto(s)...", "sub")
                pyautogui.click(self.get_posicao("btn_scroll_down"))
                time.sleep(0.4)


    # ── Relatório Excel ─────────────────────────────────────────────────────────
    def _gerar_relatorio(self, resultados=None, aba="Importação NF-e", prefixo_arquivo="relatorio_nfe"):
        if resultados is None:
            resultados = self.resultados

        if not resultados:
            self._log("⚠️  Nenhum resultado para gerar relatório.", "warn")
            return

        df = pd.DataFrame(resultados)
        wb = Workbook()
        ws = wb.active
        ws.title = aba

        hf   = Font(bold=True, color="FFFFFF")
        hfil = PatternFill("solid", fgColor="4F81BD")
        brd  = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'),  bottom=Side(style='thin'))
        aln  = Alignment(horizontal="center", vertical="center")

        for cn, col_name in enumerate(df.columns, 1):
            c = ws.cell(row=1, column=cn, value=col_name)
            c.font = hf; c.fill = hfil; c.border = brd; c.alignment = aln

        for ri, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = brd; c.alignment = aln

        for col in ws.columns:
            mx = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = mx + 2

        ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"{prefixo_arquivo}_{ts}.xlsx")
        wb.save(out)
        self._log(f"📁 Relatório salvo: {out}", "info")


if __name__ == "__main__":
    root = tk.Tk()
    app  = NFeXMLMacro(root)
    root.mainloop()